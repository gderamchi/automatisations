from __future__ import annotations

import copy
import json
import os
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from apps.workers.common.database import get_connection, init_db
from apps.workers.common.jsonio import load_json
from apps.workers.common.settings import Settings, get_settings
from apps.workers.common.templating import render_value_template, substitute_env


SUPPLY_TYPE_TO_LEDGER = {
    "carburant": ("Carburant", "Gazole"),
    "materiel": ("Fournitures", "Matériel"),
    "hotel": ("Déplacement", "Hotel/Gite"),
    "repas": ("Déplacement", "Repas"),
    "peage": ("Déplacement", "Péage autoroute"),
    "consommable": ("Fournitures", "Consommable"),
}

DOCUMENT_KIND_TO_OPERATION = {
    "invoice": "Facture fournisseur",
    "purchase_order": "Facture fournisseur",
    "receipt": "Facture fournisseur",
    "quotation": "Facture fournisseur",
    "sales_invoice": "Facture Client",
    "credit_note": "Avoir",
}

PAYMENT_STATUS_TO_LEDGER = {
    "paid": "Payé",
    "pending": "A payer",
    "unknown": "A payer",
}


def _find_mapping(mapping_name: str, settings: Settings) -> Path:
    direct = settings.excel_mappings_dir / f"{mapping_name}.json"
    example = settings.excel_mappings_dir / f"{mapping_name}.example.json"
    if direct.exists():
        return direct
    if example.exists():
        return example
    raise FileNotFoundError(f"Excel mapping not found: {mapping_name}")


def _load_document_payload(
    document_id: int,
    settings: Settings,
    *,
    document_payload_override: dict[str, Any] | None = None,
    routing_payload_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    with get_connection(settings) as connection:
        row = connection.execute(
            """
            SELECT source_name, supplier_name, supplier_siret, invoice_number, invoice_date, due_date, currency,
                   net_amount, vat_amount, gross_amount, project_ref, document_type, document_kind, supply_type,
                   payment_status, payment_date, payment_method, final_filename, validated_payload_json, normalized_payload_json
            FROM documents
            WHERE id = ?
            """,
            (document_id,),
        ).fetchone()
        routing_row = connection.execute(
            """
            SELECT corrected_payload_json, proposed_payload_json
            FROM routing_tasks
            WHERE document_id = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (document_id,),
        ).fetchone()
        project_row = None
        project_ref = row["project_ref"] if row else None
        if routing_payload_override and routing_payload_override.get("worksite_external_id"):
            project_ref = routing_payload_override["worksite_external_id"]
        elif routing_payload_override and routing_payload_override.get("target_label"):
            project_ref = routing_payload_override["target_label"]
        if project_ref:
            project_row = connection.execute(
                """
                SELECT project_code, project_name, metadata_json
                FROM doe_projects
                WHERE project_name = ? OR project_code = ? OR external_project_id = ?
                ORDER BY updated_at DESC
                LIMIT 1
                """,
                (project_ref, project_ref, project_ref),
            ).fetchone()
    if not row:
        raise KeyError(f"Document not found: {document_id}")
    payload_json = row["validated_payload_json"] or row["normalized_payload_json"]
    payload = json.loads(payload_json) if payload_json else {}
    routing_payload = json.loads((routing_row["corrected_payload_json"] or routing_row["proposed_payload_json"]) if routing_row else "{}")
    if routing_payload_override:
        routing_payload = {**routing_payload, **routing_payload_override}
    for field in (
        "source_name",
        "supplier_name",
        "supplier_siret",
        "invoice_number",
        "invoice_date",
        "due_date",
        "currency",
        "net_amount",
        "vat_amount",
        "gross_amount",
        "project_ref",
        "document_type",
        "document_kind",
        "supply_type",
        "payment_status",
        "payment_date",
        "payment_method",
        "final_filename",
    ):
        if row[field] is not None:
            payload[field] = row[field]
    if document_payload_override:
        payload.update({key: value for key, value in document_payload_override.items() if value is not None})
    for field in (
        "expense_label",
        "target_label",
        "worksite_external_id",
        "client_external_id",
        "operation_type",
        "ledger_label",
        "ledger_sub_label",
        "payment_method",
        "payment_status",
        "vat_bucket",
        "treasury_workbook_path",
        "client_ledger_path",
        "supplier_ledger_path",
        "force_type",
        "received_transfer_amount",
    ):
        if routing_payload.get(field) is not None:
            payload[field] = routing_payload[field]
    if project_row:
        if project_row["project_code"] is not None:
            payload["project_code"] = project_row["project_code"]
        if project_row["project_name"] is not None:
            payload["project_name"] = project_row["project_name"]
        metadata = json.loads(project_row["metadata_json"] or "{}")
        client = metadata.get("client") or {}
        if client.get("name"):
            payload["client_name"] = client["name"]
        if metadata.get("status"):
            payload["worksite_status"] = metadata["status"]
    payload["excel_invoice_label"] = _format_invoice_label(payload.get("invoice_number"))
    payload["excel_description"] = payload.get("expense_label") or payload.get("target_label") or payload.get("supplier_name")
    payload["excel_operation_type"] = DOCUMENT_KIND_TO_OPERATION.get(str(payload.get("document_kind") or ""), "Facture fournisseur")
    ledger_label, ledger_sub_label = SUPPLY_TYPE_TO_LEDGER.get(str(payload.get("supply_type") or ""), ("", ""))
    payload["excel_label"] = ledger_label
    payload["excel_sub_label"] = ledger_sub_label
    payload["excel_project_code"] = payload.get("project_code") or payload.get("project_ref")
    payload["excel_status_label"] = PAYMENT_STATUS_TO_LEDGER.get(str(payload.get("payment_status") or ""), "A payer")
    payload["excel_vat_rate"] = _compute_vat_rate(payload.get("net_amount"), payload.get("vat_amount"))
    payload["excel_force_type"] = ""
    payload["excel_received_transfer"] = None
    return payload


def _format_invoice_label(invoice_number: str | None) -> str | None:
    if not invoice_number:
        return None
    cleaned = str(invoice_number).strip()
    if cleaned.lower().startswith(("facture", "n°", "no", "nº")):
        return cleaned
    return f"N°{cleaned}"


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _compute_vat_rate(net_amount: Any, vat_amount: Any) -> float | None:
    net = _to_float(net_amount)
    vat = _to_float(vat_amount)
    if not net or vat is None:
        return None
    if abs(net) < 0.00001:
        return None
    return round(vat / net, 4)


def _copy_row_style(worksheet, source_row: int, target_row: int, max_column: int) -> None:
    for column in range(1, max_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)


def _copy_row_template(worksheet, source_row: int, target_row: int, max_column: int) -> None:
    _copy_row_style(worksheet, source_row, target_row, max_column)
    for column in range(1, max_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        if source_cell.data_type == "f" or (isinstance(source_cell.value, str) and source_cell.value.startswith("=")):
            target_cell.value = Translator(source_cell.value, origin=f"{get_column_letter(column)}{source_row}").translate_formula(
                f"{get_column_letter(column)}{target_row}"
            )
        else:
            target_cell.value = None


def _resolve_workbook_path(mapping: dict[str, Any], settings: Settings) -> Path:
    missing_env = [env_name for env_name in mapping.get("required_env", []) if not os.getenv(env_name)]
    if missing_env:
        raise RuntimeError(f"Missing required Excel env vars: {', '.join(missing_env)}")
    raw_path = mapping["workbook_path"].replace("${DATA_ROOT}", str(settings.data_root))
    resolved = substitute_env(raw_path).strip()
    if not resolved:
        raise RuntimeError("Excel workbook path is empty after env substitution")
    workbook_path = Path(resolved)
    if workbook_path.is_dir():
        raise RuntimeError(f"Excel workbook path points to a directory: {workbook_path}")
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    return workbook_path


def _find_table(workbook, mapping: dict[str, Any]):
    table_name = mapping["table_name"]
    if mapping.get("sheet"):
        worksheet = workbook[mapping["sheet"]]
        table = worksheet.tables.get(table_name)
        if table is None:
            raise KeyError(f"Table {table_name} not found in sheet {mapping['sheet']}")
        return worksheet, table
    for worksheet in workbook.worksheets:
        table = worksheet.tables.get(table_name)
        if table is not None:
            return worksheet, table
    raise KeyError(f"Table not found: {table_name}")


def _normalize_excel_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return datetime.fromisoformat(stripped).date()
        except ValueError:
            return value
    return value


def _write_document_to_table(document_id: int, workbook, mapping: dict[str, Any], payload: dict[str, Any]) -> tuple[str, int]:
    worksheet, table = _find_table(workbook, mapping)
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_row = min_row
    totals_row = max_row if table.totalsRowCount else None
    insert_row = totals_row or (max_row + 1)
    worksheet.insert_rows(insert_row, 1)
    source_row = int(mapping.get("template_row") or ((totals_row - 1) if totals_row else max_row))
    _copy_row_template(worksheet, source_row, insert_row, max_col)

    headers = {
        worksheet.cell(row=header_row, column=column).value: column
        for column in range(min_col, max_col + 1)
    }
    for field_name, target in mapping["columns"].items():
        column = headers.get(target)
        if column is None:
            raise KeyError(f"Column {target} not found in table {table.name}")
        rendered = payload.get(field_name)
        if rendered is None and field_name in mapping.get("constants", {}):
            rendered = render_value_template(mapping["constants"][field_name], payload)
        worksheet.cell(row=insert_row, column=column).value = _normalize_excel_value(rendered)

    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row + 1}"
    return worksheet.title, insert_row


def write_document_to_excel(document_id: int, mapping_name: str, settings: Settings | None = None) -> dict[str, Any]:
    current = settings or get_settings()
    init_db(current)
    mapping = load_json(_find_mapping(mapping_name, current))
    workbook_path = _resolve_workbook_path(mapping, current)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    payload = _load_document_payload(document_id, current)
    workbook = load_workbook(workbook_path)
    if mapping.get("table_name"):
        sheet_name, current_row = _write_document_to_table(document_id, workbook, mapping, payload)
    else:
        worksheet = workbook[mapping["sheet"]]
        sheet_name = worksheet.title
        key_column = mapping.get("key_column", "A")
        start_row = int(mapping.get("start_row", 2))
        current_row = start_row
        while worksheet[f"{key_column}{current_row}"].value not in (None, ""):
            current_row += 1

        template_row = int(mapping.get("template_row", max(start_row, current_row - 1)))
        if template_row and template_row != current_row:
            _copy_row_style(worksheet, template_row, current_row, worksheet.max_column)

        for field_name, column in mapping["columns"].items():
            rendered = payload.get(field_name)
            if rendered is None and field_name in mapping.get("constants", {}):
                rendered = render_value_template(mapping["constants"][field_name], payload)
            worksheet[f"{column}{current_row}"] = _normalize_excel_value(rendered)

    workbook.save(workbook_path)

    with get_connection(current) as connection:
        connection.execute(
            """
            UPDATE documents
            SET current_stage = 'excel_written', updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (document_id,),
        )
        connection.commit()

    return {
        "document_id": document_id,
        "mapping": mapping_name,
        "status": "success",
        "workbook_path": str(workbook_path),
        "sheet": sheet_name,
        "row": current_row,
    }


def write_document_bundle(
    document_id: int,
    mapping_names: list[str] | None = None,
    *,
    strict: bool = True,
    settings: Settings | None = None,
    document_payload_override: dict[str, Any] | None = None,
    routing_payload_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    current = settings or get_settings()
    if has_nas_excel_targets(current):
        from apps.workers.documents.nas_excel import write_nas_document_bundle

        return write_nas_document_bundle(
            document_id,
            strict=strict,
            settings=current,
            document_payload_override=document_payload_override,
            routing_payload_override=routing_payload_override,
        )

    names = list(mapping_names or list(current.default_excel_mappings))
    for mapping_name in current.optional_excel_mappings:
        if mapping_name not in names:
            names.append(mapping_name)
    results: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    for mapping_name in names:
        try:
            results.append(write_document_to_excel(document_id, mapping_name, settings=current))
        except Exception as exc:
            error = {
                "document_id": document_id,
                "mapping": mapping_name,
                "status": "error",
                "error": str(exc),
            }
            results.append(error)
            errors.append(error)
            if strict:
                raise
    if any(result.get("status") == "success" for result in results):
        with get_connection(current) as connection:
            connection.execute(
                """
                UPDATE documents
                SET current_stage = 'excel_written', updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (document_id,),
            )
            connection.commit()
    return {
        "document_id": document_id,
        "mappings": results,
        "errors": errors,
    }


def has_nas_excel_targets(settings: Settings | None = None) -> bool:
    from apps.workers.documents.nas_excel import has_nas_excel_targets as _impl

    return _impl(settings or get_settings())


def build_excel_review_payload(
    document_id: int,
    *,
    settings: Settings | None = None,
    document_payload_override: dict[str, Any] | None = None,
    routing_payload_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    from apps.workers.documents.nas_excel import build_excel_review_payload as _impl

    return _impl(
        document_id,
        settings=settings or get_settings(),
        document_payload_override=document_payload_override,
        routing_payload_override=routing_payload_override,
    )


def get_excel_form_options() -> dict[str, list[str]]:
    from apps.workers.documents.nas_excel import get_excel_form_options as _impl

    return _impl()
