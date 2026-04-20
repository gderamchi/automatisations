from __future__ import annotations

import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

from apps.workers.common.settings import Settings
from apps.workers.documents.excel import (
    _copy_row_style,
    _copy_row_template,
    _find_table,
    _format_invoice_label,
    _load_document_payload,
    _normalize_excel_value,
)


OPERATION_TYPES = [
    "Facture fournisseur",
    "Facture Client",
    "Charge salariale",
    "Charge fixe",
    "Paiement Fournisseur",
    "Paiement Client",
    "Rétrocession Engins",
    "Avoir",
]

LEDGER_OPTIONS = {
    "Carburant": ["Gazole", "Essence", "Adblue", "GNR"],
    "Fournitures": ["Matériaux", "Engin", "Sous-traitance", "Transport", "Matériel", "Petit outillage", "Consommable", "EPI", "Autre"],
    "Salariés": ["Salaire Net", "Acompte"],
    "Déplacement": ["Pension salarié", "Hotel/Gite", "Repas", "Péage autoroute", "Stationnement"],
    "Charges sociales": ["URSSAF", "PROBTP", "Mutuelle", "CIBTP", "Prévoyance", "Retraite", "Prélèvement à la source"],
    "Véhicules": ["Entretient véhicule", "Réparation véhicule", "Nettoyage véhicule", "Contrôle véhicule"],
    "Frais généraux": ["Téléphone", "Internet", "Logiciel", "Matériel informatique", "Prestation informatique", "Bureautique", "Formations", "Comptabilité", "Location bureaux", "Location Véhicule", "Sécurité"],
    "Assurance": ["Véhicule", "Responsabilté civile", "Immobilier", "Décennale"],
    "Banque": ["Frais bancaires", "Intérêts", "Emprunt", "Remboursement prêt", "Frais CB", "Crédit véhicule"],
    "Client": ["Facturation", "Avoir", "Retenue de garantie", "Rétrocession engins"],
}
PAYMENT_METHODS = [
    "Virement",
    "Prélévement",
    "Chéque",
    "Espéce",
    "CB-5148-BP",
    "CB-2289-BP",
    "CB-5198-FORTUNEO",
    "CB-4033-N26",
    "CB-4649-QONTOF",
    "CB-8097-QONTOA",
    "CB-8551-QONTOS",
]
PAYMENT_STATUSES = ["A payer", "Payé", "En attente", "En retard"]
VAT_BUCKETS = ["0", "5.5", "10", "20"]


CLIENT_LEDGER_HEADERS = [
    "Date",
    "N° Facture",
    "ID Document",
    "TTC",
    "Échéance",
    "Débit (+)",
    "Crédit (-)",
    "Solde",
    "Date paiement",
    "Description",
    "Force type",
    "Type d'opération",
    "Fournisseur",
    "Mode paiement",
    "Libellé",
    "Sous-Libellé",
    "N° Chantier",
    "HT",
    "% TVA 5,5",
    "% TVA 10",
    "% TVA 20",
    "Déductible",
    "Collectée",
    "Statut",
    "Virement Reçu",
    "Fournisseur Code",
    "Final",
]

TREASURY_HEADERS = [
    "Validation",
    "DATE",
    "TRANSACTION",
    "CATÉGORIE",
    "MONTANT",
    "€",
    "DESCRIPTION / LIBELLE / REFERENCE RELEVE BANCAIRE",
    "ID - Document",
    "N° Facture",
    "Type paiement",
    "Mode Paiement",
    "Fournisseur",
    "Libellé",
    "Sous-libellé",
    "N° Chantier affecté",
    "Ht",
    "TTC",
    "% TVA 5,5",
    "% TVA 10",
    "% TVA 20",
    "TVA Déductible",
    "TVA collecté",
]

SUPPLIER_LEDGER_HEADERS = [
    "Date",
    "N° Facture",
    "Description",
    "N° Chantier",
    "HT",
    "% TVA",
    "TVA",
    "TTC",
    "Échéance",
    "Débit (+)",
    "Crédit (-)",
    "Solde",
    "Date paiement",
    "Mode paiement",
    "Statut",
]


def has_nas_excel_targets(settings: Settings) -> bool:
    return any(
        (
            settings.accounting_share_root,
            settings.resolved_client_ledgers_root,
            settings.resolved_supplier_ledgers_root,
            settings.resolved_treasury_ledgers_root,
        )
    )


def get_excel_form_options() -> dict[str, list[str]]:
    sub_labels = sorted({value for values in LEDGER_OPTIONS.values() for value in values})
    return {
        "operation_types": OPERATION_TYPES,
        "ledger_labels": list(LEDGER_OPTIONS.keys()),
        "ledger_sub_labels": sub_labels,
        "payment_methods": PAYMENT_METHODS,
        "payment_statuses": PAYMENT_STATUSES,
        "vat_buckets": VAT_BUCKETS,
    }


def _normalize_token(value: str | None) -> str:
    raw = unicodedata.normalize("NFKD", str(value or ""))
    stripped = "".join(char for char in raw if not unicodedata.combining(char))
    cleaned = []
    for char in stripped.lower():
        cleaned.append(char if char.isalnum() else " ")
    return " ".join("".join(cleaned).split())


def _coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        try:
            return datetime.fromisoformat(value.strip()).date()
        except ValueError:
            return None
    return None


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def _vat_bucket_from_amounts(payload: dict[str, Any], explicit: str | None) -> str:
    if explicit in {"0", "5.5", "10", "20"}:
        return explicit
    net_amount = _to_float(payload.get("net_amount"))
    vat_amount = _to_float(payload.get("vat_amount"))
    if not net_amount or vat_amount is None:
        return "0"
    rate = round((vat_amount / net_amount) * 100, 1)
    if abs(rate - 5.5) < 0.3:
        return "5.5"
    if abs(rate - 10.0) < 0.3:
        return "10"
    if abs(rate - 20.0) < 0.3:
        return "20"
    return "0"


def _is_client_operation(operation_type: str) -> bool:
    lowered = operation_type.lower()
    return "client" in lowered and "paiement fournisseur" not in lowered


def _derive_operation_type(payload: dict[str, Any], routing_payload: dict[str, Any]) -> str:
    explicit = routing_payload.get("operation_type")
    if explicit:
        return explicit
    if payload.get("document_kind") == "sales_invoice":
        return "Facture Client"
    if payload.get("document_kind") == "credit_note":
        return "Avoir"
    return payload.get("excel_operation_type") or "Facture fournisseur"


def _derive_ledger_label(payload: dict[str, Any], routing_payload: dict[str, Any], operation_type: str) -> str:
    explicit = routing_payload.get("ledger_label")
    if explicit:
        return explicit
    if _is_client_operation(operation_type):
        return "Client"
    return payload.get("excel_label") or "Fournitures"


def _derive_ledger_sub_label(payload: dict[str, Any], routing_payload: dict[str, Any], ledger_label: str) -> str:
    explicit = routing_payload.get("ledger_sub_label")
    if explicit:
        return explicit
    if ledger_label in LEDGER_OPTIONS and payload.get("excel_sub_label") in LEDGER_OPTIONS[ledger_label]:
        return payload["excel_sub_label"]
    return LEDGER_OPTIONS.get(ledger_label, ["Autre"])[0]


def _derive_payment_status(payload: dict[str, Any], routing_payload: dict[str, Any]) -> str:
    explicit = routing_payload.get("payment_status")
    if explicit:
        return explicit
    return payload.get("excel_status_label") or "A payer"


def _derive_payment_method(payload: dict[str, Any], routing_payload: dict[str, Any]) -> str:
    explicit = routing_payload.get("payment_method")
    if explicit:
        return explicit
    return payload.get("payment_method") or "Virement"


def _list_workbooks(root: Path | None) -> list[Path]:
    if not root or not root.exists():
        return []
    return sorted(
        path
        for path in root.rglob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$") and "#recycle" not in path.parts
    )


def _resolve_workbook_by_token(root: Path | None, token: str | None) -> Path | None:
    workbooks = _list_workbooks(root)
    if not workbooks:
        return None
    if not token:
        return workbooks[0] if len(workbooks) == 1 else None

    normalized = _normalize_token(token)
    matches = [path for path in workbooks if normalized and normalized in _normalize_token(path.stem)]
    if not matches:
        return None
    matches.sort(key=lambda path: (len(_normalize_token(path.stem)), path.name))
    return matches[0]


def _exercise_folder_for(invoice_date: date) -> str:
    start_year = invoice_date.year if invoice_date.month >= 10 else invoice_date.year - 1
    return f"{start_year}_{start_year + 1}"


def _resolve_treasury_workbook(root: Path | None, invoice_date: date | None, explicit_path: str | None) -> tuple[Path | None, list[dict[str, str]]]:
    workbooks = _list_workbooks(root)
    options = [{"label": path.name, "value": str(path)} for path in workbooks]
    if explicit_path:
        explicit = Path(explicit_path)
        return (explicit if explicit.exists() else None), options
    if not root or not invoice_date:
        return None, options

    exercise_folder = _exercise_folder_for(invoice_date)
    prefix = f"{invoice_date.year}-{invoice_date.month:02d}-01_FINANCE_TRESORERIE_"
    for path in workbooks:
        if exercise_folder in path.parts and path.name.startswith(prefix):
            return path, options
    return None, options


def _resolve_client_ledger(payload: dict[str, Any], settings: Settings, explicit_path: str | None) -> Path | None:
    if explicit_path:
        explicit = Path(explicit_path)
        return explicit if explicit.exists() else None
    client_name = payload.get("client_name")
    return _resolve_workbook_by_token(settings.resolved_client_ledgers_root, client_name)


def _resolve_supplier_ledger(payload: dict[str, Any], settings: Settings, explicit_path: str | None) -> Path | None:
    if explicit_path:
        explicit = Path(explicit_path)
        return explicit if explicit.exists() else None
    supplier_name = payload.get("supplier_name")
    return _resolve_workbook_by_token(settings.resolved_supplier_ledgers_root, supplier_name)


def _build_target_options(
    workbooks: list[Path],
    *,
    recommended_path: Path | None,
    selected_path: Path | None,
) -> list[dict[str, Any]]:
    options: list[dict[str, Any]] = []
    seen: set[str] = set()
    listed_values = {str(path) for path in workbooks}

    def add_option(path: Path | None, *, recommended: bool = False, manual: bool = False) -> None:
        if not path:
            return
        value = str(path)
        if value in seen:
            return
        label = path.name
        if recommended:
            label = f"{label} (Recommandé)"
        elif manual:
            label = f"{label} (Chemin manuel)"
        options.append({"label": label, "value": value, "recommended": recommended})
        seen.add(value)

    add_option(recommended_path, recommended=True)
    add_option(
        selected_path,
        manual=selected_path is not None and str(selected_path) not in listed_values,
    )
    for path in workbooks:
        add_option(path, recommended=path == recommended_path)
    return options


def _build_target_state(
    *,
    label: str,
    workbooks: list[Path],
    selected_path: Path | None,
    recommended_path: Path | None,
    invalid_value: str | None,
    missing_message: str,
) -> dict[str, Any]:
    options = _build_target_options(
        workbooks,
        recommended_path=recommended_path,
        selected_path=selected_path,
    )
    if selected_path:
        auto_recovered = bool(
            invalid_value
            and recommended_path
            and str(selected_path) == str(recommended_path)
            and invalid_value != str(selected_path)
        )
        return {
            "label": label,
            "path": str(selected_path),
            "status": "ready",
            "message": "Le fichier recommandé a été préselectionné." if auto_recovered else "Fichier prêt pour l'écriture.",
            "error": None,
            "options": options,
            "selected_value": str(selected_path),
            "recommended_value": str(recommended_path) if recommended_path else "",
        }
    if options:
        return {
            "label": label,
            "path": None,
            "status": "needs_selection",
            "message": (
                "Le chemin précédent est introuvable. Choisissez un fichier."
                if invalid_value
                else "Choisissez un fichier pour continuer."
            ),
            "error": None,
            "options": options,
            "selected_value": "",
            "recommended_value": str(recommended_path) if recommended_path else "",
        }
    return {
        "label": label,
        "path": None,
        "status": "missing",
        "message": missing_message,
        "error": missing_message,
        "options": [],
        "selected_value": "",
        "recommended_value": "",
    }


def _resolve_rate_columns(vat_bucket: str) -> tuple[float, float, float]:
    if vat_bucket == "5.5":
        return 0.055, 0.0, 0.0
    if vat_bucket == "10":
        return 0.0, 0.10, 0.0
    if vat_bucket == "20":
        return 0.0, 0.0, 0.20
    return 0.0, 0.0, 0.0


def _build_review_values(payload: dict[str, Any], routing_payload: dict[str, Any]) -> dict[str, Any]:
    operation_type = _derive_operation_type(payload, routing_payload)
    ledger_label = _derive_ledger_label(payload, routing_payload, operation_type)
    ledger_sub_label = _derive_ledger_sub_label(payload, routing_payload, ledger_label)
    payment_status = _derive_payment_status(payload, routing_payload)
    payment_method = _derive_payment_method(payload, routing_payload)
    vat_bucket = _vat_bucket_from_amounts(payload, routing_payload.get("vat_bucket"))
    vat_5_5, vat_10, vat_20 = _resolve_rate_columns(vat_bucket)
    vat_amount = _to_float(payload.get("vat_amount")) or 0.0
    is_client = _is_client_operation(operation_type)
    gross_amount = _to_float(payload.get("gross_amount"))
    net_amount = _to_float(payload.get("net_amount"))
    due_date = _coerce_date(payload.get("due_date"))
    invoice_date = _coerce_date(payload.get("invoice_date"))
    payment_date = _coerce_date(payload.get("payment_date"))
    force_type = routing_payload.get("force_type") or ""
    received_transfer_amount = _to_float(routing_payload.get("received_transfer_amount")) or None

    return {
        "invoice_date": invoice_date,
        "invoice_label": _format_invoice_label(payload.get("invoice_number")),
        "document_id_label": payload.get("final_filename") or f"document-{payload.get('id')}",
        "gross_amount": gross_amount,
        "due_date": due_date,
        "payment_date": payment_date,
        "description": routing_payload.get("expense_label") or payload.get("excel_description") or payload.get("supplier_name"),
        "force_type": force_type,
        "operation_type": operation_type,
        "supplier_name": payload.get("supplier_name"),
        "payment_method": payment_method,
        "ledger_label": ledger_label,
        "ledger_sub_label": ledger_sub_label,
        "project_code": payload.get("project_code") or payload.get("project_ref"),
        "net_amount": net_amount,
        "vat_bucket": vat_bucket,
        "vat_rate_5_5": vat_5_5,
        "vat_rate_10": vat_10,
        "vat_rate_20": vat_20,
        "vat_deductible": vat_amount if not is_client else 0.0,
        "vat_collectee": vat_amount if is_client else 0.0,
        "payment_status": payment_status,
        "received_transfer_amount": received_transfer_amount,
        "supplier_code": payload.get("supplier_siret"),
        "final_label": force_type or operation_type,
        "transaction_label": operation_type,
        "category_label": routing_payload.get("expense_label") or ledger_sub_label,
        "currency": payload.get("currency") or "EUR",
    }


def build_excel_review_payload(
    document_id: int,
    *,
    settings: Settings,
    document_payload_override: dict[str, Any] | None = None,
    routing_payload_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = _load_document_payload(
        document_id,
        settings,
        document_payload_override=document_payload_override,
        routing_payload_override=routing_payload_override,
    )
    payload["id"] = document_id
    routing_payload = routing_payload_override or {}
    values = _build_review_values(payload, routing_payload)
    treasury_selected_path, treasury_options_raw = _resolve_treasury_workbook(
        settings.resolved_treasury_ledgers_root,
        values["invoice_date"],
        routing_payload.get("treasury_workbook_path"),
    )
    treasury_recommended_path, _ = _resolve_treasury_workbook(
        settings.resolved_treasury_ledgers_root,
        values["invoice_date"],
        None,
    )
    treasury_workbooks = [Path(option["value"]) for option in treasury_options_raw]
    treasury_active_path = treasury_selected_path or treasury_recommended_path

    client_selected_path = _resolve_client_ledger(payload, settings, routing_payload.get("client_ledger_path"))
    client_recommended_path = _resolve_client_ledger(payload, settings, None)
    client_workbooks = _list_workbooks(settings.resolved_client_ledgers_root)
    client_active_path = client_selected_path or client_recommended_path

    supplier_selected_path = _resolve_supplier_ledger(payload, settings, routing_payload.get("supplier_ledger_path"))
    supplier_recommended_path = _resolve_supplier_ledger(payload, settings, None)
    supplier_workbooks = _list_workbooks(settings.resolved_supplier_ledgers_root)
    supplier_active_path = supplier_selected_path or supplier_recommended_path

    targets: dict[str, dict[str, Any]] = {}
    errors: list[str] = []

    if has_nas_excel_targets(settings):
        targets["treasury"] = _build_target_state(
            label="Suivi trésorerie",
            workbooks=treasury_workbooks,
            selected_path=treasury_active_path,
            recommended_path=treasury_recommended_path,
            invalid_value=routing_payload.get("treasury_workbook_path"),
            missing_message="Classeur de trésorerie introuvable",
        ) | {
            "table_name": "Tableau8",
            "sheet_name": None,
            "values": [
                ("Validation", "Validé"),
                ("DATE", values["invoice_date"]),
                ("TRANSACTION", values["transaction_label"]),
                ("CATÉGORIE", values["category_label"]),
                ("MONTANT", values["gross_amount"]),
                ("€", values["currency"]),
                ("DESCRIPTION / LIBELLE / REFERENCE RELEVE BANCAIRE", values["description"]),
                ("ID - Document", values["document_id_label"]),
                ("N° Facture", values["invoice_label"]),
                ("Type paiement", values["operation_type"]),
                ("Mode Paiement", values["payment_method"]),
                ("Fournisseur", values["supplier_name"]),
                ("Libellé", values["ledger_label"]),
                ("Sous-libellé", values["ledger_sub_label"]),
                ("N° Chantier affecté", values["project_code"]),
                ("Ht", values["net_amount"]),
                ("TTC", values["gross_amount"]),
                ("% TVA 5,5", values["vat_rate_5_5"]),
                ("% TVA 10", values["vat_rate_10"]),
                ("% TVA 20", values["vat_rate_20"]),
                ("TVA Déductible", values["vat_deductible"]),
                ("TVA collecté", values["vat_collectee"]),
            ],
        }
        targets["client"] = _build_target_state(
            label="Grand livre client",
            workbooks=client_workbooks,
            selected_path=client_active_path,
            recommended_path=client_recommended_path,
            invalid_value=routing_payload.get("client_ledger_path"),
            missing_message="Grand livre client introuvable",
        ) | {
            "table_name": "Tresorerie",
            "sheet_name": None,
            "values": [
                ("Date", values["invoice_date"]),
                ("N° Facture", values["invoice_label"]),
                ("ID Document", values["document_id_label"]),
                ("TTC", values["gross_amount"]),
                ("Échéance", values["due_date"]),
                ("Date paiement", values["payment_date"]),
                ("Description", values["description"]),
                ("Force type", values["force_type"]),
                ("Type d'opération", values["operation_type"]),
                ("Fournisseur", values["supplier_name"]),
                ("Mode paiement", values["payment_method"]),
                ("Libellé", values["ledger_label"]),
                ("Sous-Libellé", values["ledger_sub_label"]),
                ("N° Chantier", values["project_code"]),
                ("HT", values["net_amount"]),
                ("% TVA 5,5", values["vat_rate_5_5"]),
                ("% TVA 10", values["vat_rate_10"]),
                ("% TVA 20", values["vat_rate_20"]),
                ("Déductible", values["vat_deductible"]),
                ("Collectée", values["vat_collectee"]),
                ("Statut", values["payment_status"]),
                ("Virement Reçu", values["received_transfer_amount"]),
                ("Fournisseur Code", values["supplier_code"]),
                ("Final", values["final_label"]),
            ],
        }
        targets["supplier"] = _build_target_state(
            label="Grand livre fournisseur",
            workbooks=supplier_workbooks,
            selected_path=supplier_active_path,
            recommended_path=supplier_recommended_path,
            invalid_value=routing_payload.get("supplier_ledger_path"),
            missing_message="Grand livre fournisseur introuvable",
        ) | {
            "values": [
                ("Date", values["invoice_date"]),
                ("N° Facture", values["invoice_label"]),
                ("Description", values["description"]),
                ("N° Chantier", values["project_code"]),
                ("HT", values["net_amount"]),
                ("% TVA", (values["vat_rate_5_5"] or values["vat_rate_10"] or values["vat_rate_20"])),
                ("TVA", values["vat_deductible"]),
                ("TTC", values["gross_amount"]),
                ("Échéance", values["due_date"]),
                ("Date paiement", values["payment_date"]),
                ("Mode paiement", values["payment_method"]),
                ("Statut", values["payment_status"]),
            ],
        }
        errors = [target["error"] for target in targets.values() if target.get("error")]

    defaults = {
        "operation_type": values["operation_type"],
        "ledger_label": values["ledger_label"],
        "ledger_sub_label": values["ledger_sub_label"],
        "payment_method": values["payment_method"],
        "payment_status": values["payment_status"],
        "vat_bucket": values["vat_bucket"],
        "treasury_workbook_path": str(treasury_active_path) if treasury_active_path else "",
        "client_ledger_path": str(client_active_path) if client_active_path else "",
        "supplier_ledger_path": str(supplier_active_path) if supplier_active_path else "",
        "force_type": values["force_type"],
        "received_transfer_amount": values["received_transfer_amount"] or "",
    }
    return {
        "enabled": has_nas_excel_targets(settings),
        "errors": errors,
        "targets": targets,
        "defaults": defaults,
    }


def _target_value_map(target: dict[str, Any]) -> dict[str, Any]:
    return {header: value for header, value in target["values"]}


def _write_values_to_table(workbook_path: Path, table_name: str, values: dict[str, Any], preferred_sheet: str | None = None) -> tuple[str, int]:
    workbook = load_workbook(workbook_path)
    worksheet, table = _find_table(workbook, {"sheet": preferred_sheet, "table_name": table_name} if preferred_sheet else {"table_name": table_name})
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    totals_row = max_row if table.totalsRowCount else None
    insert_row = totals_row or (max_row + 1)
    worksheet.insert_rows(insert_row, 1)
    source_row = max(min_row + 1, (totals_row - 1) if totals_row else max_row)
    _copy_row_template(worksheet, source_row, insert_row, max_col)

    headers = {
        worksheet.cell(row=min_row, column=column).value: column
        for column in range(min_col, max_col + 1)
    }
    for header, value in values.items():
        column = headers.get(header)
        if column is None:
            continue
        worksheet.cell(row=insert_row, column=column).value = _normalize_excel_value(value)

    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row + 1}"
    workbook.save(workbook_path)
    return worksheet.title, insert_row


def _find_supplier_header_row(worksheet) -> int:
    for row_index in range(1, min(25, worksheet.max_row) + 1):
        row_values = [worksheet.cell(row=row_index, column=column).value for column in range(1, worksheet.max_column + 1)]
        if "Date" in row_values and "N° Facture" in row_values and "Description" in row_values:
            return row_index
    raise KeyError(f"Supplier ledger header row not found in {worksheet.title}")


def _write_supplier_rows(workbook_path: Path, values: dict[str, Any]) -> tuple[str, int]:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = _find_supplier_header_row(worksheet)
    headers = {
        worksheet.cell(row=header_row, column=column).value: column
        for column in range(1, worksheet.max_column + 1)
    }
    data_start = header_row + 1
    insert_row = data_start
    while worksheet.cell(row=insert_row, column=1).value not in (None, ""):
        insert_row += 1

    create_payment_row = values.get("payment_status") == "Payé" and values.get("payment_date")
    rows_to_insert = 2 if create_payment_row else 1
    worksheet.insert_rows(insert_row, rows_to_insert)

    invoice_template_row = data_start
    payment_template_row = min(data_start + 1, worksheet.max_row)
    _copy_row_template(worksheet, invoice_template_row, insert_row, worksheet.max_column)
    if create_payment_row:
        _copy_row_template(worksheet, payment_template_row, insert_row + 1, worksheet.max_column)

    invoice_values = {
        "Date": values.get("Date"),
        "N° Facture": values.get("N° Facture"),
        "Description": "Facture",
        "N° Chantier": values.get("N° Chantier"),
        "HT": values.get("HT"),
        "% TVA": values.get("% TVA"),
        "TVA": values.get("TVA"),
        "TTC": values.get("TTC"),
        "Échéance": values.get("Échéance"),
        "Date paiement": values.get("Date paiement"),
        "Mode paiement": values.get("Mode paiement"),
        "Statut": values.get("Statut"),
    }
    for header, value in invoice_values.items():
        column = headers.get(header)
        if column is None:
            continue
        worksheet.cell(row=insert_row, column=column).value = _normalize_excel_value(value)

    if create_payment_row:
        payment_values = {
            "Date": values.get("Date paiement"),
            "Description": "Paiement",
            "Débit (+)": values.get("TTC"),
            "Date paiement": values.get("Date paiement"),
            "Mode paiement": values.get("Mode paiement"),
            "Statut": values.get("Statut"),
        }
        for header, value in payment_values.items():
            column = headers.get(header)
            if column is None:
                continue
            worksheet.cell(row=insert_row + 1, column=column).value = _normalize_excel_value(value)

    workbook.save(workbook_path)
    return worksheet.title, insert_row


def write_nas_document_bundle(
    document_id: int,
    *,
    strict: bool,
    settings: Settings,
    document_payload_override: dict[str, Any] | None = None,
    routing_payload_override: dict[str, Any] | None = None,
) -> dict[str, Any]:
    review = build_excel_review_payload(
        document_id,
        settings=settings,
        document_payload_override=document_payload_override,
        routing_payload_override=routing_payload_override,
    )
    results: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for target_key in ("treasury", "client", "supplier"):
        target = review["targets"].get(target_key)
        if not target:
            continue
        if target["status"] != "ready" or not target.get("path"):
            error = {
                "document_id": document_id,
                "mapping": target_key,
                "status": "error",
                "error": target.get("error") or target.get("message") or "Excel target unavailable",
            }
            results.append(error)
            errors.append(error)
            if strict:
                raise RuntimeError(error["error"])
            continue

        workbook_path = Path(target["path"])
        values = _target_value_map(target)
        try:
            if target_key == "supplier":
                sheet_name, row_number = _write_supplier_rows(workbook_path, values)
            else:
                sheet_name, row_number = _write_values_to_table(
                    workbook_path,
                    target["table_name"],
                    values,
                    preferred_sheet=target.get("sheet_name"),
                )
            results.append(
                {
                    "document_id": document_id,
                    "mapping": target_key,
                    "status": "success",
                    "workbook_path": str(workbook_path),
                    "sheet": sheet_name,
                    "row": row_number,
                }
            )
        except Exception as exc:
            error = {
                "document_id": document_id,
                "mapping": target_key,
                "status": "error",
                "error": str(exc),
            }
            results.append(error)
            errors.append(error)
            if strict:
                raise

    return {
        "document_id": document_id,
        "mappings": results,
        "errors": errors,
        "review": review,
    }
