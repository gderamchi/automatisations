from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.workers.common.database import get_connection
from apps.workers.common.settings import get_settings
from apps.workers.documents.excel import build_excel_review_payload, write_document_bundle
from apps.workers.documents.ingest import ingest_document
from apps.workers.documents.ocr_service import run_document_ocr
from apps.workers.doe.service import upsert_project


CLIENT_TABLE_HEADERS = [
    "Date",
    "N° Facture",
    "ID Document",
    "TTC",
    "Échéance",
    "Débit (+)",
    "Crédit (-)",
    "Solde",
    "Date paiement",
    "Description",
    "Force type",
    "Type d'opération",
    "Fournisseur",
    "Mode paiement",
    "Libellé",
    "Sous-Libellé",
    "N° Chantier",
    "HT",
    "% TVA 5,5",
    "% TVA 10",
    "% TVA 20",
    "Déductible",
    "Collectée",
    "Statut",
    "Virement Reçu",
    "Fournisseur Code",
    "Final",
]


def _create_client_ledger_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Suivie-tresorerie_2026_04"
    for column, header in enumerate(CLIENT_TABLE_HEADERS, start=1):
        worksheet.cell(row=15, column=column, value=header)
    worksheet["A17"] = date(2026, 1, 1)
    worksheet["B17"] = "N°EXISTANT"
    worksheet["C17"] = "existing.pdf"
    worksheet["D17"] = 100
    worksheet["F17"] = '=IF(Z17="Facture Client",D17,"")'
    worksheet["G17"] = '=IF(Z17="Paiement Client",-Y17,"")'
    worksheet["H17"] = '=IF(AND(F17="",G17=""),"",ROUND(IF(H16="",0,H16)+IF(F17="",0,F17)+IF(G17="",0,G17),2))'
    worksheet["Q17"] = "CCM-001"
    worksheet["R17"] = 100
    worksheet["S17"] = 0
    worksheet["T17"] = 0
    worksheet["U17"] = 0.2
    worksheet["V17"] = 20
    worksheet["W17"] = 0
    worksheet["X17"] = "A payer"
    worksheet["AA17"] = '=IF(AND(K17="",L17=""),"",IF(K17<>"",K17,L17))'
    worksheet["A18"] = "Total"
    table = Table(displayName="Tresorerie", ref="A15:AA18")
    table.totalsRowCount = 1
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)
    workbook.save(path)


def _create_treasury_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Avril"
    headers = [
        "Validation",
        "DATE",
        "TRANSACTION",
        "CATÉGORIE",
        "MONTANT",
        "€",
        "DESCRIPTION / LIBELLE / REFERENCE RELEVE BANCAIRE",
        "ID - Document",
        "N° Facture",
        "Type paiement",
        "Mode Paiement",
        "Fournisseur",
        "Libellé",
        "Sous-libellé",
        "N° Chantier affecté",
        "Ht",
        "TTC",
        "% TVA 5,5",
        "% TVA 10",
        "% TVA 20",
        "TVA Déductible",
        "TVA collecté",
    ]
    for index, header in enumerate(headers, start=39):
        worksheet.cell(row=1, column=index, value=header)
    worksheet["AN2"] = date(2026, 4, 1)
    worksheet["AO2"] = "Facture fournisseur"
    worksheet["AQ2"] = 10
    worksheet["AR2"] = "€"
    worksheet["AS2"] = "Template"
    worksheet["AT2"] = "template-id"
    worksheet["AV2"] = "Paiement Fournisseur"
    worksheet["AW2"] = "Virement"
    worksheet["AX2"] = "JORISIDE"
    worksheet["AY2"] = "Fournitures"
    worksheet["AZ2"] = "Matériaux"
    worksheet["BA2"] = "CCM-001"
    worksheet["BB2"] = 10
    worksheet["BC2"] = 12
    worksheet["BF2"] = 0.2
    worksheet["BG2"] = 2
    table = Table(displayName="Tableau8", ref="AM1:BH2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)
    workbook.save(path)


def _create_supplier_ledger_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Feuil1"
    headers = [
        "Date",
        "N° Facture",
        "Description",
        "N° Chantier",
        "HT",
        "% TVA",
        "TVA",
        "TTC",
        "Échéance",
        "Débit (+)",
        "Crédit (-)",
        "Solde",
        "Date paiement",
        "Mode paiement",
        "Statut",
    ]
    for index, header in enumerate(headers, start=1):
        worksheet.cell(row=8, column=index, value=header)
    worksheet["A9"] = date(2026, 4, 1)
    worksheet["B9"] = "FAC-1"
    worksheet["C9"] = "Facture"
    worksheet["D9"] = "CCM-001"
    worksheet["E9"] = 100
    worksheet["F9"] = 0.2
    worksheet["G9"] = 20
    worksheet["H9"] = 120
    worksheet["I9"] = date(2026, 4, 30)
    worksheet["J9"] = 0
    worksheet["K9"] = "=H9"
    worksheet["L9"] = "=J9-K9"
    worksheet["O9"] = "Payé"
    worksheet["A10"] = date(2026, 4, 30)
    worksheet["C10"] = "Paiement"
    worksheet["J10"] = 120
    worksheet["L10"] = "=L9+J10-K10"
    worksheet["M10"] = date(2026, 4, 30)
    worksheet["N10"] = "Virement"
    worksheet["O10"] = "Payé"
    workbook.save(path)


@pytest.fixture()
def nas_excel_root(tmp_path, monkeypatch):
    share_root = tmp_path / "Professionnel_CCM" / "02_Comptabilité_Vert"
    client_root = share_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_CLIENT"
    supplier_root = share_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_FOURNISSEUR"
    treasury_root = share_root / "04_EXPERT_COMPTABLE" / "06_Suivie_Tresorerie" / "2025_2026"
    client_root.mkdir(parents=True, exist_ok=True)
    supplier_root.mkdir(parents=True, exist_ok=True)
    treasury_root.mkdir(parents=True, exist_ok=True)

    _create_client_ledger_workbook(client_root / "CLI-0001_FRAMPAS_GRAND_LIVRE_CLIENT.xlsx")
    _create_supplier_ledger_workbook(supplier_root / "FOU-0001_JORISIDE_GRAND_LIVRE_FOURNISSEUR.xlsx")
    _create_treasury_workbook(treasury_root / "2026-04-01_FINANCE_TRESORERIE_AVRIL_2026_en cours.xlsx")

    monkeypatch.setenv("ACCOUNTING_SHARE_ROOT", str(share_root))
    get_settings.cache_clear()
    yield share_root
    get_settings.cache_clear()


def _prepare_document(test_settings, tmp_path, sample_invoice_text: str, *, supplier_name: str = "JORISIDE"):
    invoice = tmp_path / "invoice.txt"
    invoice.write_text(sample_invoice_text, encoding="utf-8")
    ingested = ingest_document(str(invoice), "manual", settings=test_settings)
    run_document_ocr(ingested["document_id"], settings=test_settings)
    with get_connection(test_settings) as connection:
        connection.execute(
            """
            UPDATE documents
            SET supplier_name = ?,
                document_kind = 'invoice',
                supply_type = 'materiel',
                invoice_number = 'FAC-2026-001',
                invoice_date = '2026-04-12',
                due_date = '2026-04-30',
                net_amount = '1000.00',
                vat_amount = '200.00',
                gross_amount = '1200.00',
                final_filename = '2026-04-12_facture_joriside_ccm-002_1200.00.pdf'
            WHERE id = ?
            """,
            (supplier_name, ingested["document_id"]),
        )
        connection.commit()
    upsert_project(
        external_project_id="53032",
        project_code="CCM-002",
        project_name="12576-6-RONDEAU - CCM2026-002",
        metadata={"client": {"id": 1950319, "name": "FRAMPAS"}, "status": "signed"},
        settings=test_settings,
    )
    return ingested["document_id"]


def test_build_excel_review_payload_resolves_nas_targets(nas_excel_root, test_settings, tmp_path, sample_invoice_text):
    document_id = _prepare_document(test_settings, tmp_path, sample_invoice_text)
    review = build_excel_review_payload(
        document_id,
        settings=test_settings,
        routing_payload_override={
            "target_label": "12576-6-RONDEAU - CCM2026-002",
            "worksite_external_id": "53032",
        },
    )

    assert review["enabled"] is True
    assert review["errors"] == []
    assert review["targets"]["treasury"]["status"] == "ready"
    assert review["targets"]["client"]["status"] == "ready"
    assert review["targets"]["supplier"]["status"] == "ready"
    assert review["defaults"]["treasury_workbook_path"].endswith("2026-04-01_FINANCE_TRESORERIE_AVRIL_2026_en cours.xlsx")
    assert review["defaults"]["client_ledger_path"].endswith("CLI-0001_FRAMPAS_GRAND_LIVRE_CLIENT.xlsx")
    assert review["defaults"]["supplier_ledger_path"].endswith("FOU-0001_JORISIDE_GRAND_LIVRE_FOURNISSEUR.xlsx")
    assert review["targets"]["treasury"]["recommended_value"].endswith("2026-04-01_FINANCE_TRESORERIE_AVRIL_2026_en cours.xlsx")
    assert review["targets"]["client"]["options"][0]["recommended"] is True
    assert review["targets"]["supplier"]["options"][0]["recommended"] is True


def test_build_excel_review_payload_marks_selectable_targets_as_needs_selection(nas_excel_root, test_settings, tmp_path, sample_invoice_text):
    extra_client_ledger = nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_CLIENT" / "CLI-0002_AUTRE_GRAND_LIVRE_CLIENT.xlsx"
    _create_client_ledger_workbook(extra_client_ledger)
    document_id = _prepare_document(test_settings, tmp_path, sample_invoice_text, supplier_name="INCONNU")
    review = build_excel_review_payload(document_id, settings=test_settings)

    assert review["enabled"] is True
    assert review["errors"] == []
    assert review["targets"]["treasury"]["status"] == "ready"
    assert review["targets"]["client"]["status"] == "needs_selection"
    assert review["targets"]["client"]["options"]
    assert review["targets"]["client"]["message"] == "Choisissez un fichier pour continuer."
    assert review["targets"]["supplier"]["status"] == "needs_selection"
    assert review["targets"]["supplier"]["options"]
    assert review["targets"]["supplier"]["message"] == "Choisissez un fichier pour continuer."


def test_write_document_bundle_writes_all_nas_targets(nas_excel_root, test_settings, tmp_path, sample_invoice_text):
    document_id = _prepare_document(test_settings, tmp_path, sample_invoice_text)
    result = write_document_bundle(
        document_id,
        strict=True,
        settings=test_settings,
        routing_payload_override={
            "target_label": "12576-6-RONDEAU - CCM2026-002",
            "worksite_external_id": "53032",
            "operation_type": "Facture fournisseur",
            "ledger_label": "Fournitures",
            "ledger_sub_label": "Matériaux",
            "payment_method": "Virement",
            "payment_status": "A payer",
            "vat_bucket": "20",
            "final_filename": "2026-04-12_facture_joriside_ccm-002_1200.00.pdf",
        },
    )

    assert result["errors"] == []
    client_book = load_workbook(
        nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_CLIENT" / "CLI-0001_FRAMPAS_GRAND_LIVRE_CLIENT.xlsx"
    )
    treasury_book = load_workbook(
        nas_excel_root / "04_EXPERT_COMPTABLE" / "06_Suivie_Tresorerie" / "2025_2026" / "2026-04-01_FINANCE_TRESORERIE_AVRIL_2026_en cours.xlsx"
    )
    supplier_book = load_workbook(
        nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_FOURNISSEUR" / "FOU-0001_JORISIDE_GRAND_LIVRE_FOURNISSEUR.xlsx"
    )

    client_sheet = client_book["Suivie-tresorerie_2026_04"]
    assert client_sheet["A18"].value.date() == date(2026, 4, 12)
    assert client_sheet["B18"].value == "N°FAC-2026-001"
    assert client_sheet["M18"].value == "JORISIDE"
    assert client_sheet["O18"].value == "Fournitures"

    treasury_sheet = treasury_book["Avril"]
    assert treasury_sheet["AN3"].value.date() == date(2026, 4, 12)
    assert treasury_sheet["AT3"].value == "2026-04-12_facture_joriside_ccm-002_1200.00.pdf"
    assert treasury_sheet["AX3"].value == "JORISIDE"
    assert treasury_sheet["AY3"].value == "Fournitures"

    supplier_sheet = supplier_book["Feuil1"]
    assert supplier_sheet["A11"].value.date() == date(2026, 4, 12)
    assert supplier_sheet["B11"].value == "N°FAC-2026-001"
    assert supplier_sheet["D11"].value == "CCM-002"
    assert supplier_sheet["O11"].value == "A payer"


def test_routing_submit_accepts_selected_supplier_and_client_ledgers(client, nas_excel_root, test_settings, tmp_path, incomplete_invoice_text):
    invoice = tmp_path / "invoice-selected-ledgers.txt"
    invoice.write_text(incomplete_invoice_text, encoding="utf-8")

    ingest = client.post(
        "/internal/documents/ingest",
        headers={"X-Internal-Token": "test-token"},
        json={"source_path": str(invoice), "source_kind": "manual"},
    )
    document_id = ingest.json()["document_id"]
    ocr = client.post(f"/internal/documents/{document_id}/ocr", headers={"X-Internal-Token": "test-token"})
    validation_token = ocr.json()["validation_token"]

    client.post(
        f"/validate/{validation_token}",
        auth=("validator", "secret"),
        data={
            "document_type": "purchase_invoice",
            "document_kind": "invoice",
            "decision": "approve",
            "validator_name": "ops",
            "supply_type": "materiel",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-777",
            "invoice_date": "2026-04-10",
            "due_date": "2026-04-30",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "project_ref": "PROJET-X",
        },
    )

    upsert_project(
        external_project_id="53032",
        project_code="CCM-002",
        project_name="12576-6-RONDEAU - CCM2026-002",
        metadata={"client": {"id": 1950319, "name": "FRAMPAS"}, "status": "signed"},
        settings=test_settings,
    )
    with get_connection(test_settings) as connection:
        routing_token = connection.execute(
            "SELECT token FROM routing_tasks WHERE document_id = ? ORDER BY id DESC LIMIT 1",
            (document_id,),
        ).fetchone()["token"]

    client_ledger_path = nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_CLIENT" / "CLI-0001_FRAMPAS_GRAND_LIVRE_CLIENT.xlsx"
    supplier_ledger_path = nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_FOURNISSEUR" / "FOU-0001_JORISIDE_GRAND_LIVRE_FOURNISSEUR.xlsx"

    response = client.post(
        f"/route/{routing_token}",
        auth=("validator", "secret"),
        data={
            "decision": "approve",
            "validator_name": "ops",
            "document_kind": "invoice",
            "supply_type": "materiel",
            "expense_label": "Achat de matériel INCONNU",
            "routing_confidence": "0.94",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-777",
            "invoice_date": "2026-04-10",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "worksite_external_id": "53032",
            "operation_type": "Facture fournisseur",
            "ledger_label": "Fournitures",
            "ledger_sub_label": "Matériaux",
            "payment_method": "Virement",
            "payment_status": "A payer",
            "vat_bucket": "20",
            "client_ledger_path_selected": str(client_ledger_path),
            "supplier_ledger_path_selected": str(supplier_ledger_path),
        },
    )

    assert response.status_code == 200
    assert "Dispatch: dispatched" in response.text


def test_routing_submit_manual_supplier_ledger_overrides_selected_value(client, nas_excel_root, test_settings, tmp_path, incomplete_invoice_text):
    invoice = tmp_path / "invoice-manual-ledger.txt"
    invoice.write_text(incomplete_invoice_text, encoding="utf-8")

    ingest = client.post(
        "/internal/documents/ingest",
        headers={"X-Internal-Token": "test-token"},
        json={"source_path": str(invoice), "source_kind": "manual"},
    )
    document_id = ingest.json()["document_id"]
    ocr = client.post(f"/internal/documents/{document_id}/ocr", headers={"X-Internal-Token": "test-token"})
    validation_token = ocr.json()["validation_token"]

    client.post(
        f"/validate/{validation_token}",
        auth=("validator", "secret"),
        data={
            "document_type": "purchase_invoice",
            "document_kind": "invoice",
            "decision": "approve",
            "validator_name": "ops",
            "supply_type": "materiel",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-778",
            "invoice_date": "2026-04-10",
            "due_date": "2026-04-30",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "project_ref": "PROJET-X",
        },
    )

    upsert_project(
        external_project_id="53032",
        project_code="CCM-002",
        project_name="12576-6-RONDEAU - CCM2026-002",
        metadata={"client": {"id": 1950319, "name": "FRAMPAS"}, "status": "signed"},
        settings=test_settings,
    )
    with get_connection(test_settings) as connection:
        routing_token = connection.execute(
            "SELECT token FROM routing_tasks WHERE document_id = ? ORDER BY id DESC LIMIT 1",
            (document_id,),
        ).fetchone()["token"]

    selected_supplier_ledger = nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_FOURNISSEUR" / "FOU-0001_JORISIDE_GRAND_LIVRE_FOURNISSEUR.xlsx"
    manual_supplier_ledger = tmp_path / "MANUAL_SUPPLIER_LEDGER.xlsx"
    _create_supplier_ledger_workbook(manual_supplier_ledger)

    response = client.post(
        f"/route/{routing_token}",
        auth=("validator", "secret"),
        data={
            "decision": "approve",
            "validator_name": "ops",
            "document_kind": "invoice",
            "supply_type": "materiel",
            "expense_label": "Achat de matériel INCONNU",
            "routing_confidence": "0.94",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-778",
            "invoice_date": "2026-04-10",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "worksite_external_id": "53032",
            "operation_type": "Facture fournisseur",
            "ledger_label": "Fournitures",
            "ledger_sub_label": "Matériaux",
            "payment_method": "Virement",
            "payment_status": "A payer",
            "vat_bucket": "20",
            "supplier_ledger_path_selected": str(selected_supplier_ledger),
            "supplier_ledger_path_manual": str(manual_supplier_ledger),
        },
    )

    assert response.status_code == 200
    assert "Dispatch: dispatched" in response.text

    manual_book = load_workbook(manual_supplier_ledger)
    selected_book = load_workbook(selected_supplier_ledger)
    assert manual_book["Feuil1"]["A11"].value.date() == date(2026, 4, 10)
    assert selected_book["Feuil1"]["A11"].value is None


def test_routing_submit_blocks_when_supplier_ledger_missing(client, nas_excel_root, test_settings, tmp_path, incomplete_invoice_text):
    invoice = tmp_path / "invoice-missing-supplier.txt"
    invoice.write_text(incomplete_invoice_text, encoding="utf-8")

    ingest = client.post(
        "/internal/documents/ingest",
        headers={"X-Internal-Token": "test-token"},
        json={"source_path": str(invoice), "source_kind": "manual"},
    )
    document_id = ingest.json()["document_id"]
    ocr = client.post(f"/internal/documents/{document_id}/ocr", headers={"X-Internal-Token": "test-token"})
    validation_token = ocr.json()["validation_token"]

    client.post(
        f"/validate/{validation_token}",
        auth=("validator", "secret"),
        data={
            "document_type": "purchase_invoice",
            "document_kind": "invoice",
            "decision": "approve",
            "validator_name": "ops",
            "supply_type": "materiel",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-404",
            "invoice_date": "2026-04-10",
            "due_date": "2026-04-30",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "project_ref": "PROJET-X",
        },
    )

    upsert_project(
        external_project_id="53032",
        project_code="CCM-002",
        project_name="12576-6-RONDEAU - CCM2026-002",
        metadata={"client": {"id": 1950319, "name": "FRAMPAS"}, "status": "signed"},
        settings=test_settings,
    )
    supplier_root = nas_excel_root / "04_EXPERT_COMPTABLE" / "05_Grand_Livre" / "GRAND_LIVRE_FOURNISSEUR"
    for workbook in supplier_root.glob("*.xlsx"):
        workbook.unlink()
    with get_connection(test_settings) as connection:
        routing_token = connection.execute(
            "SELECT token FROM routing_tasks WHERE document_id = ? ORDER BY id DESC LIMIT 1",
            (document_id,),
        ).fetchone()["token"]

    response = client.post(
        f"/route/{routing_token}",
        auth=("validator", "secret"),
        data={
            "decision": "approve",
            "validator_name": "ops",
            "document_kind": "invoice",
            "supply_type": "materiel",
            "expense_label": "Achat de matériel INCONNU",
            "routing_confidence": "0.94",
            "supplier_name": "INCONNU",
            "invoice_number": "FAC-2026-404",
            "invoice_date": "2026-04-10",
            "currency": "EUR",
            "net_amount": "1000.00",
            "vat_amount": "200.00",
            "gross_amount": "1200.00",
            "worksite_external_id": "53032",
            "operation_type": "Facture fournisseur",
            "ledger_label": "Fournitures",
            "ledger_sub_label": "Matériaux",
            "payment_method": "Virement",
            "payment_status": "A payer",
            "vat_bucket": "20",
        },
    )

    assert response.status_code == 200
    assert "Validation bloquée" in response.text
    assert "Grand livre fournisseur introuvable" in response.text
    with get_connection(test_settings) as connection:
        routing_status = connection.execute(
            "SELECT status FROM routing_tasks WHERE token = ?",
            (routing_token,),
        ).fetchone()["status"]
    assert routing_status == "pending"
