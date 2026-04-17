from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.workers.common.database import get_connection
from apps.workers.documents.excel import write_document_bundle
from apps.workers.documents.ingest import ingest_document
from apps.workers.documents.ocr_service import run_document_ocr


CLIENT_TABLE_HEADERS = [
    "Date",
    "N° Facture",
    "ID Document",
    "TTC",
    "Échéance",
    "Débit (+)",
    "Crédit (-)",
    "Solde",
    "Date paiement",
    "Description",
    "Force type",
    "Type d'opération",
    "Fournisseur",
    "Mode paiement",
    "Libellé",
    "Sous-Libellé",
    "N° Chantier",
    "HT",
    "% TVA",
    "Déductible",
    "Collectée",
    "Statut",
    "Virement Reçu",
    "Fournisseur Code",
    "Final",
]


def _create_client_ledger_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Suivie-tresorerie_2026_04"
    for column, header in enumerate(CLIENT_TABLE_HEADERS, start=1):
        worksheet.cell(row=15, column=column, value=header)
    worksheet["A17"] = date(2026, 1, 1)
    worksheet["B17"] = "N°EXISTANT"
    worksheet["C17"] = "existing.pdf"
    worksheet["D17"] = '=SUM(Tresorerie[[#This Row],[HT]],Tresorerie[[#This Row],[Déductible]],Tresorerie[[#This Row],[Collectée]])'
    worksheet["F17"] = '=IF(Tresorerie[[#This Row],[Final]]="Facture Client",D17,"")'
    worksheet["G17"] = '=IF(Tresorerie[[#This Row],[Final]]="Paiement Client",-Tresorerie[[#This Row],[Virement Reçu]],IF(Tresorerie[[#This Row],[Final]]="Rétrocession Engins",-Tresorerie[[#This Row],[TTC]],""))'
    worksheet["H17"] = '=IF(AND(F17="",G17=""),"",ROUND(IF(H16="",0,H16)+IF(F17="",0,F17)+IF(G17="",0,G17),2))'
    worksheet["J17"] = "Existing description"
    worksheet["L17"] = "Facture fournisseur"
    worksheet["M17"] = "Existing supplier"
    worksheet["O17"] = "Fournitures"
    worksheet["P17"] = "Matériel"
    worksheet["R17"] = 100
    worksheet["S17"] = 0.2
    worksheet["T17"] = '=IF(Y17="Rétrocession Engins",R17*S17,"")'
    worksheet["U17"] = '=IF(Y17="Facture client",R17*S17,"")'
    worksheet["V17"] = "A payer"
    worksheet["Y17"] = '=IF(AND(K17="",L17=""),"",IF(K17<>"",K17,L17))'
    worksheet["A18"] = "Total"
    table = Table(displayName="Tresorerie", ref="A15:Y18")
    table.totalsRowCount = 1
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium1", showRowStripes=True, showColumnStripes=False)
    worksheet.add_table(table)
    workbook.save(path)


def test_write_document_bundle_supports_client_ledger_table(tmp_path, test_settings, sample_invoice_text, monkeypatch):
    workbook_path = tmp_path / "client-ledger.xlsx"
    _create_client_ledger_workbook(workbook_path)
    monkeypatch.setenv("CLIENT_GRAND_LIVRE_WORKBOOK_PATH", str(workbook_path))

    invoice = tmp_path / "invoice.txt"
    invoice.write_text(sample_invoice_text, encoding="utf-8")
    ingested = ingest_document(str(invoice), "manual", settings=test_settings)
    run_document_ocr(ingested["document_id"], settings=test_settings)
    with get_connection(test_settings) as connection:
        connection.execute(
            """
            UPDATE documents
            SET document_kind = 'invoice',
                supply_type = 'materiel',
                project_ref = 'CCM-002',
                final_filename = '2026-03-10_achat_materiel_amazon_ccm-002_1200.00.pdf'
            WHERE id = ?
            """,
            (ingested["document_id"],),
        )
        connection.commit()

    result = write_document_bundle(
        ingested["document_id"],
        mapping_names=["client_grand_livre"],
        strict=False,
        settings=test_settings,
    )

    assert result["errors"] == []
    workbook = load_workbook(workbook_path)
    worksheet = workbook["Suivie-tresorerie_2026_04"]
    assert worksheet.tables["Tresorerie"].ref == "A15:Y19"
    assert worksheet["A18"].value.date() == date(2026, 3, 10)
    assert worksheet["B18"].value == "N°FAC-2026-001"
    assert worksheet["C18"].value == "2026-03-10_achat_materiel_amazon_ccm-002_1200.00.pdf"
    assert worksheet["J18"].value == "Dépense ACME BTP SAS"
    assert worksheet["L18"].value == "Facture fournisseur"
    assert worksheet["M18"].value == "ACME BTP SAS"
    assert worksheet["O18"].value == "Fournitures"
    assert worksheet["P18"].value == "Matériel"
    assert worksheet["Q18"].value == "CCM-002"
    assert worksheet["R18"].value == "1000.00"
    assert worksheet["S18"].value == 0.2
    assert worksheet["V18"].value == "A payer"
